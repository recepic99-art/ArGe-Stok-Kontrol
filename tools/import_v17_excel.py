"""Eski Excel stok ve hareket sayfalarını web demosunun JSON yapısına aktarır."""

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
OUTPUT_FILE = PROJECT_DIR / "arge-numune-depo.json"
SEED_FILE = PROJECT_DIR / "assets" / "js" / "seed.js"


def text(value):
    """Excel'deki boş değerleri JSON için güvenli metne çevirir."""
    return "" if value is None else str(value)


def number(value):
    """Miktar alanlarını sayı olarak saklar."""
    if value in (None, ""):
        return 0
    numeric = float(value)
    return int(numeric) if numeric.is_integer() else numeric


def date_text(value):
    """Excel tarihlerini uygulamanın kullandığı okunabilir biçime çevirir."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y %H:%M")
    raw = text(value).strip()
    for pattern in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(raw, pattern).strftime("%d.%m.%Y %H:%M")
        except ValueError:
            pass
    return raw


def rows_by_header(sheet):
    """İlk satırdaki başlıkları anahtar olarak kullanıp diğer satırları sözlüğe çevirir."""
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value).strip() for value in next(rows)]
    return [
        dict(zip(headers, values))
        for values in rows
        if any(value not in (None, "") for value in values)
    ]


def convert_excel(excel_file):
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    stock_rows = rows_by_header(workbook["Stok"])
    movement_rows = rows_by_header(workbook["Hareketler"])

    items = [
        {
            "id": text(row["ID"]),
            "name": text(row["Malzeme Adi"]),
            "category": text(row["Cins/Kategori"]),
            "footprint": text(row.get("Kilif")),
            "box": text(row["Kutu"]),
            "quantity": number(row["Miktar"]),
            "unit": text(row["Birim"]) or "adet",
            "critical": number(row["Kritik Seviye"]),
            "description": text(row["Aciklama"]),
            "updatedAt": date_text(row["Son Guncelleme"]),
        }
        for row in stock_rows
    ]

    logs = [
        {
            "id": text(row["Islem No"]) or f"log-{index}",
            "tableId": "t-main",
            "itemId": text(row["Malzeme ID"]),
            "itemName": text(row["Malzeme Adi"]),
            "date": date_text(row["Tarih Saat"]),
            "type": text(row["Islem Tipi"]),
            "quantity": number(row["Miktar"]),
            "purpose": text(row["Kullanim Amaci/Proje"]),
            "user": text(row["Alan/Veren"]),
            "note": text(row["Aciklama"]),
            "before": number(row["Onceki Stok"]),
            "after": number(row["Yeni Stok"]),
        }
        for index, row in enumerate(movement_rows, start=1)
    ]

    data = {
        "schemaVersion": 2,
        "users": [
            {
                "id": "u-recep",
                "username": "recep",
                "name": "Recep İç",
                "password": "demo123",
            }
        ],
        "tables": [
            {
                "id": "t-main",
                "name": "Numune Malzemeler Listesi",
                "items": items,
            }
        ],
        "logs": logs,
        "session": {
            "currentUserId": None,
            "activeTableId": "t-main",
            "openTableIds": ["t-main"],
        },
        "settings": {
            "theme": "light",
            "leftWidth": 250,
            "rightWidth": 330,
            "historyHeight": 38,
            "leftCollapsed": False,
            "rightCollapsed": False,
            "visibleColumns": [
                "id",
                "name",
                "category",
                "footprint",
                "box",
                "quantity",
                "unit",
                "critical",
                "description",
                "updatedAt",
            ],
            "columnOrder": [
                "id",
                "name",
                "category",
                "footprint",
                "box",
                "quantity",
                "unit",
                "critical",
                "description",
                "updatedAt",
            ],
        },
    }

    OUTPUT_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    SEED_FILE.write_text(
        "window.DepoSeedState = " +
        json.dumps(data, ensure_ascii=False, indent=2) +
        ";\n",
        encoding="utf-8",
    )
    return len(items), len(logs)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("excel_file", type=Path)
    arguments = parser.parse_args()
    item_count, log_count = convert_excel(arguments.excel_file)
    print(f"{item_count} stok kartı ve {log_count} hareket aktarıldı.")
